import collections
import openpyxl
import pandas as pd

MONTH_COLLECTION = (
    {1: 'January',
     2: 'February',
     3: 'March',
     4: 'April',
     5: 'May',
     6: 'June',
     7: 'July',
     8: 'August',
     9: 'September',
     10: 'October',
     11: 'November',
     12: 'December'}
)


def make_report(log_file_name, report_template_file_name, report_output_file_name):
    """
    :param log_file_name: имя файла логов
    :param report_template_file: имя файла-шаблона для отчета
    :param report_output_file_name: имя файла отчета
    :return: создает отчет в формате report.xlsx
    """
    # Чтение и анализ данных из файла excel
    df = pd.read_excel(log_file_name, sheet_name='log')

    # Считаем количество сессий по браузерам, сортируем по месяцам и запаковываем в словарь. Считаем самые популярные браузеры.
    session_counter = collections.Counter(df['Браузер'])
    sorted_month_sessions = df.groupby(df['Дата посещения'].dt.strftime('%B'))['Браузер'].value_counts()
    sms_dict = sorted_month_sessions.to_dict()

    # Сортируем и суммируем строки купленных товаров по всем покупкам по месяцам. Считаем самые продаваемые товары
    sorted_month_goods = df.groupby(df['Дата посещения'].dt.strftime('%B'))['Купленные товары'].apply(','.join)
    smg_dict = sorted_month_goods.to_dict()
    sales_counter = collections.Counter((sum([val.strip().split(',') for val in smg_dict.values()], [])))

    # Собираем все покупки в разрезе пола в словарь и суммируем их
    goods_dict = df.groupby(df['Пол'])['Купленные товары'].apply(','.join)

    # Открываем файл шаблона отчета report_template.xlsx
    wb = openpyxl.load_workbook(filename=report_template_file_name, data_only=True)
    ws = wb.active

    # Выполняем запись данных в объект wb
    # Цикл для 1-го задания:
    browser = 0  # Номер браузера в самых популярных браузерах
    month_num = 1
    for row in range(5, 12):  # Строки на листе шаблона
        month_num = 1  # Столбцы на листе шаблона
        for column in range(2, 14):
            try:
                ws.cell(row=row, column=column, value=sms_dict[
                    MONTH_COLLECTION[month_num], session_counter.most_common(7)[browser][
                        0]])  # Заполняем значения использования
                ws.cell(row=row, column=1,
                        value=session_counter.most_common(7)[browser][0])  # Заполняем названия браузеров

            except:
                ws.cell(row=row, column=column, value=0)

            ws.cell(row=12, column=column, value=f'=SUM({chr(column + 64)}5:{chr(column + 64)}11)')
            month_num += 1
        browser += 1

    # Цикл для 2-го задания:
    good = 0  # Номер товара в самых популярных товарах
    month_num = 1
    for row in range(19, 26):  # Строки на листе шаблона
        month_num = 1
        for column in range(2, 14):  # Столбцы на листе шаблона

            try:
                ws.cell(row=row, column=column,
                        value=collections.Counter(smg_dict[MONTH_COLLECTION[month_num]].split(','))[
                            sales_counter.most_common(7)[good][0]])  # Заполняем значения проданных товаров
                ws.cell(row=row, column=1, value=sales_counter.most_common(7)[good][0])  # Заполняем названия товаров

            except:
                ws.cell(row=row, column=column, value=0)

            ws.cell(row=26, column=column, value=f'=SUM({chr(column + 64)}19:{chr(column + 64)}25)')
            month_num += 1
        good += 1

        # Цикл для 3-го задания
    women_goods_list = goods_dict['ж']  # Выделяем женские покупки из словаря
    most_popular_women_goods = collections.Counter(
        women_goods_list.strip().split(',')).most_common()  # Считаем популярные покупки
    most_popular_women_good = most_popular_women_goods[0][0]  # Берем самый популярный товар
    most_unpopular_women_good = most_popular_women_goods[-1][0]  # Берем самый непопулярный товар

    men_goods_list = goods_dict['м']  # Выделяем мужские покупки из словаря
    most_popular_men_goods = collections.Counter(
        men_goods_list.strip().split(',')).most_common()  # Считаем популярные покупки
    most_popular_men_good = most_popular_men_goods[0][0]  # Берем самый популярный товар
    most_unpopular_men_good = most_popular_men_goods[-1][0]  # Берем самый непопулярный товар

    # Записываем полученные данные в шаблон
    ws.cell(row=31, column=2, value=most_popular_men_good)
    ws.cell(row=32, column=2, value=most_popular_women_good)
    ws.cell(row=33, column=2, value=most_unpopular_men_good)
    ws.cell(row=34, column=2, value=most_unpopular_women_good)

    # Сохраняем файл отчета
    wb.save(report_output_file_name)
